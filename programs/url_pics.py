"""
Programa para bajar imagenes de una liga proveniente de un arechivo de excel
columna C de excel debe tener numero de parte
columna k de excel debe contener la liga URL
archivo de excel se debe llamar  "Registros Ingram/ligas_foto.xlsx")
"""

import openpyxl
import pandas as pd
from urllib import request, error
from time import time
from datetime import datetime
import ssl
import certifi
from pathlib import Path

# create a secure SSL context using certifi CA bundle
CTX = ssl.create_default_context(cafile=certifi.where())

def count_elapsed_time(f):
    """
    Decorator.
    Execute the function and calculate the elapsed time.
    Print the result to the standard output.
    """
    def wrapper(*args, **kwargs):
        start_time = time()
        ret = f(*args, **kwargs)
        elapsed_time = time() - start_time
        print("Elapsed time: %0.10f seconds." % elapsed_time)
        return ret
    return wrapper

@count_elapsed_time
def read_excel(result, errors) -> list:
    item_dict = {}
    excel_path = Path("programs/static/files_to_read/ligas_foto.xlsx")

    #try to open excel file
    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        errors.append([None, None, str(excel_path), "Input Excel file not found"])
        print(f"ERROR: input file not found: {excel_path}")
        return
    except Exception as e:
        errors.append([None, None, str(excel_path), f"Failed to open Excel file: {e}"])
        print(f"ERROR: failed to open Excel file: {excel_path}, error: {e}")
        return
    
    sheet = wb.active

    item_index = 0
    # start from row 2 to skip header
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        record = [item_index] + list(row)

        # protect access to record[5] using try/except instead of len-check
        try:
            item_key = record[5]
        except IndexError:
            errors.append([item_index, None, None, "row too short"])
            item_index += 1
            continue

        # check for duplicate items (assumes item number at position 5)
        if item_key in item_dict:
            errors.append([item_index, item_key, record[2] if len(record) > 2 else None,
                           f"duplicate item, original line: {item_dict[item_key]}"])
            item_index += 1
            continue
        item_dict[item_key] = record[2] if len(record) > 2 else None

        result.append(record)
        item_index += 1
    return

@count_elapsed_time
def import_picts(result, errors) -> None:
    cont = 0
    out_dir = Path("programs/static/url_images")
    out_dir.mkdir(parents=True, exist_ok=True)

    for i, rec in enumerate(result):
        try:
            remote_url = rec[13]    # index 13 is "URL"
        except IndexError:
            errors.append([rec[0], None, None, "missing URL index"])
            continue

        local_file = out_dir / f"{rec[5]}.jpg"

        try:
            with request.urlopen(remote_url, context=CTX, timeout=30) as resp:
                data = resp.read()
            local_file.write_bytes(data)
            print(i, remote_url, str(local_file))
            cont += 1
        except (error.HTTPError, error.URLError, ValueError, TimeoutError) as e:
            errors.append([rec[2] if len(rec) > 2 else None, rec[5] if len(rec) > 5 else None, remote_url, f"download failed: {e}"])
        except Exception as e:
            errors.append([rec[2] if len(rec) > 2 else None, rec[5] if len(rec) > 5 else None, remote_url, f"unexpected error: {e}"])

    print(cont, "Archivos descargados")
    print(len(errors), "Errores encontrados")

@count_elapsed_time
def print_errors(init, errors) -> None:
    HEADERS = ["Excel Line", "Item Number", "URL", "Error Description"]   # column names for the error report
    datos = pd.DataFrame(errors, columns=HEADERS)
    out_file = Path("programs/static/errores_excel")
    out_file.mkdir(parents=True, exist_ok=True)
    excel_path = out_file / f"errores_{init}.xlsx"
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        datos.to_excel(writer, sheet_name="Hoja 1", index=False)
    print("Reporte generado:", excel_path)


# main
now = datetime.now()
init = now.strftime('day_%d_%m_%Y_time_%H_%M')
result = []
errors = []
read_excel(result, errors)

if len(errors) == 0:    # proceed to download only if no errors in reading excel
    import_picts(result, errors)
else:
    print(f"Skipping downloads: {len(errors)} error(s) found during read_excel")

if not errors:
    print("No errors to report")
else:
    print_errors(init, errors)