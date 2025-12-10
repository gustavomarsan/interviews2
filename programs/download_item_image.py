"""
Programa para bajar imagenes de internet con numero de ITEM sin ligas
descraga la primera imagen de internet y genera un archivo de excel con 100 ligas por item

"""

from serpapi import GoogleSearch
import openpyxl
import pandas as pd
from urllib import request, response
from time import time
from datetime import datetime
import ssl
import os
import aiohttp
import asyncio
from pathlib import Path
ssl._create_default_https_context = ssl._create_unverified_context


# https://serpapi.com/    to put in line 42  "api_key"  = 

def count_elapsed_time(f):
    """
    Decorator.
    Execute the function and calculate the elapsed time.
    Print the result to the standard output.
    """
    def wrapper(*args, **kwargs):
        # Start counting.
        start_time = time()
        # Take the original function's return value.
        ret = f(*args, **kwargs)
        # Calculate the elapsed time.
        elapsed_time = time() - start_time
        print("Elapsed time: %0.10f seconds." % elapsed_time)
        return ret
    return wrapper

async def descargar_primera_imagen(item, urls_search, errors, i, api_key, session):        # SerpApi
    item_urls = [item]      # set all the urls of the search
    params = {
        "q": item,
        "tbm": "isch",      # Modo imágenes
        "api_key": api_key   # key from SerpApi 
    }
    print(i, params)

    resultados = GoogleSearch(params).get_dict()

    if "images_results" not in resultados:
        errors.append(item)
        return False
    
    try :
        url_imagen = resultados["images_results"][0]["original"]
        print(url_imagen)    
        headers = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}
        async with session.get(url_imagen, headers=headers, ssl=False) as response:
            data = await response.read()
            content_type = response.headers.get("Content-Type", "")
            
            # Debug logs
            print("Status:", response.status)
            print("Content-Type:", content_type)

            is_jpeg = data[:3] == b'\xff\xd8\xff'
            is_png  = data[:8] == b'\x89PNG\r\n\x1a\n'
            is_gif  = data[:3] == b'GIF'            

            # Validate the content before writing the file
            is_image = ("image" in content_type or is_jpeg or is_png or is_gif)

            if response.status == 200 and is_image:
                image_path = Path("programs/static/photos_api") / f"{item}.jpg"
                with open(image_path, "wb") as handler:
                    handler.write(data)
            else:
                print(f"❌ {item}: URL is not an image or request failed.")
                errors.append(item)
                return False

        # store all the urls found
        for j in range(len(resultados["images_results"])) :
            item_urls.append(resultados["images_results"][j]["original"])
        
        urls_search.append(list(item_urls))
        return True
        
    except Exception as e:
        print(f"❌ ERROR {item}: {e}")
        errors.append(item)
        return False
        

#@count_elapsed_time
def read_excel(result) -> None :
    # Define variable to load the dataframe
    excel_path = Path("programs/static/files_to_read/items.xlsx")
    dataframe = openpyxl.load_workbook(excel_path)

    # Define variable to read sheet
    dataframe1 = dataframe.active

    # Iterate the loop to read the cell values
    index = 0
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(0, dataframe1.max_column):
            result.append(str(col[row].value))
        index += 1
    print("Archivo Excel leido")
    return

@count_elapsed_time
async def descargar_imgenes_lista(result, errors, urls_search, init) -> None :
    API_KEY = os.getenv("SERPAPI_API_KEY")
    if not API_KEY:
        raise RuntimeError("SERPAPI_API_KEY not set in environment")

    async with aiohttp.ClientSession() as session:
        tasks = []
        for i in range(len(result)):
            item = result[i]
            tasks.append(descargar_primera_imagen(item, urls_search, errors, i, API_KEY, session))

        await asyncio.gather(*tasks)


    print((len(result)-len(errors)),"Imagenes descargadas")
    print(len(errors), "items no encontrados")
    
    # Save all the urls found to an excel file
    results_path = Path("programs/static/results_excel/")
    datos = pd.DataFrame(urls_search)
    excel_writer = pd.ExcelWriter(results_path / (init + ".xlsx"))
    datos.to_excel(excel_writer, sheet_name=init)
    excel_writer._save()


@count_elapsed_time
def print_errors(errors, init)-> None :
    datos = pd.DataFrame(errors)
    results_path = Path("programs/static/results_excel/")
    excel_writer = pd.ExcelWriter(results_path / ("errores_api_"+init+".xlsx"))
    datos.to_excel(excel_writer, sheet_name=init)
    excel_writer._save()
    print("Reporte de errores generado")



now = datetime.now()
print("prueba", now)
init = str(now.strftime('day_%d_%m_%Y_time_%H_%M'))
result = []
errors = []
urls_search = []        #set a list of urls found for each item for future use
read_excel(result)
print(result)

if len(result) > 0 :
    asyncio.run(descargar_imgenes_lista(result, errors, urls_search, init))

if len(errors) > 0 :
    print_errors(errors, init)